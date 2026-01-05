import json
import logging
import re
import tiktoken
import asyncio
from flask import Blueprint, request, jsonify, Response, stream_with_context, current_app
from together import Together
from auth import optional_token_required
from memory import TokenAwareMemoryManager
from db import get_db_connection, get_unauthorized_request_count, increment_unauthorized_request_count, return_db_connection
from routes.together_key_routes import decrypt_key
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import requests
from datetime import datetime, timezone
import os

# Import tool execution
from tools import execute_tool, format_tool_result_for_llm

chat_bp = Blueprint('chat_bp', __name__)

THINK_TAG_REGEX = re.compile(r'<think>.*?</think>', re.DOTALL)

def current_date():
    return datetime.now(timezone.utc).astimezone().strftime("%A, %B %d, %Y")

BASE_SYSTEM_PROMPT = """
# Core Instructions (DO NOT OVERRIDE)
You are Deepthinks, a context-aware AI assistant.

## Memory System
- **LONG-TERM MEMORY**: Appears as "Here is a summary of the conversation so far:".
- **SHORT-TERM MEMORY**: The most recent user/assistant message exchanges, provided for immediate context.

## Response Style & Verbosity Control
Your response length must be guided by the complexity and nature of the user's request:
1.  **Low Verbosity (Direct):** For simple, factual, or direct questions (e.g., "What is 2+2?", "Define gravity."), respond **concisely** (1-2 sentences) and get straight to the answer.
2.  **High Verbosity (Detailed):** For complex questions, requests for analysis, comparison, code, or detailed explanation, provide a **thorough and detailed** response.
3.  **Efficiency:** Always prioritize the most token-efficient output while meeting the verbosity requirement.

## Tool Access

**Available Tools:**
- search_web: Current information (news, weather, stocks, etc.)
- email_tool: Gmail operations (search, read, send, manage emails) - **FULLY AGENTIC**

**Decision Rule:**
Explicit request ΓåÆ Execute immediately
Uncertain ΓåÆ Ask first

**Format:**
Write natural text, END with: {{"tool_call": "tool_name", "query": "brief task description"}}

**Examples:**
Γ£ô "Bitcoin price today?" ΓåÆ "I'll check the latest market data. {{"tool_call": "search_web", "query": "current bitcoin price USD today"}}"
Γ£ô "Find email from John" ΓåÆ "I'll search your emails. {{"tool_call": "email_tool", "query": "find latest email from John"}}"
Γ£ô "Read emails from Sarah in SENT and reply to her" ΓåÆ "I'll handle that for you. {{"tool_call": "email_tool", "query": "find emails from Sarah in SENT folder and send her a reply"}}"
Γ£ô "AI history?" ΓåÆ "I can overview. Want 2024-2025 specifics?" (No JSON - waiting for confirmation)

**Rules:**
- One tool at a time
- JSON must be LAST
- Never output JSON when just offering

**CRITICAL: email_tool is FULLY AGENTIC:**
1. **ONE CALL FOR ENTIRE TASK**: Always provide the COMPLETE multi-step task in a SINGLE tool call. NEVER split into multiple calls.
   Γ¥î BAD: First call "search emails from John" ΓåÆ then second call "send reply to John"
   Γ£à GOOD: Single call "find emails from John and send him a reply saying [brief intent]"
2. **CONCISE QUERIES ONLY**: State WHAT you want done, not HOW to do it. email_tool decides its own execution strategy.
   Γ¥î BAD: "Search emails using from_addr parameter with value john@email.com in SENT label"  
   Γ£à GOOD: "find emails from John in SENT folder"
3. **NO EMAIL BODY COMPOSITION**: NEVER include full email content/body in the query. Just state the intent.
   Γ¥î BAD: "send email to sarah@co.com with subject 'Meeting' and body 'Hi Sarah, I wanted to follow up on our conversation...'"
   Γ£à GOOD: "send sarah@co.com an email following up on our meeting"
4. **NO CONTEXT INJECTION**: DO NOT include conversation history or context in the query. email_tool can access chat history if needed.
   Γ¥î BAD: "Based on our discussion about the Python project, send email to john saying..."
   Γ£à GOOD: "send email to john about the Python project urging him to hurry up"
5. **TRUST THE AGENT**: email_tool will ask for missing info (like email addresses) on its own. Don't try to solve everything in the query.


## Important Guidelines
1.  **Prioritize Memory**: Use the long-term and short-term memory.
2.  **Trust Recent Information**: If recent user messages contradict long-term memory, the most recent information takes precedence.
3.  **Be Context-Aware**: Do not explicitly mention your memory system as its proprietary. Use the context it provides to have natural, informed conversations.
4.  **Clarifying Questions**: If ambiguous, ask clarifying questions before solving
5. **Coding Requirement**: - If the user has a coding-related request, recommend using the Deepcode feature. This mode leverages the most powerful open-source coding model available.
- Prompt the user to enable Deepcode by toggling the Deepcode switch in the app.
- When Deepcode is enabled, memory will automatically switch to JSON format (this indicates the mode change, when you see JSON response in conversation history that's mean user just used code mode and now it's turned off. Note: Do NOT repeat the same JSON structure as now the deepcode mode is turned off.).

# current date. 
Current Date is {today}. Always take Reference of this date for any time related scenarios.
# User Information
The user's preferred name is: {user_name}

# User-Defined Persona
**User-defined persona**: Use this user-defined persona for shaping your Tone and behavior requested by user.
{user_persona}
"""

# Code mode system prompt with tool support
CODE_SYSTEM_PROMPT_TEMPLATE = """
# Core Instructions (DO NOT OVERRIDE)
You are Deepthinks, a context-aware AI assistant.

## Memory System
- **LONG-TERM MEMORY**: Appears as "Here is a summary of the conversation so far:" containing:
    - `interactions`: An array of past conversation summaries, verbatim context, and priority scores (0-10).
    - `important_details`: A list of key facts, user preferences, and persistent information.
- **SHORT-TERM MEMORY**: The most recent user/assistant message exchanges, provided for immediate context.

## Tool Architecture in Code Mode
You respond ONLY in JSON format following the CodeResponse schema. Tool calls are integrated at specific points in your response flow:

**Available Tools:**
- search_web: Current information (news, weather, stocks, etc.)
- email_tool: Gmail operations (search, read, send, manage emails)

### Available Tool Call Points:
1. **tool_after_text**: After writing the Text field (before Files)
2. **tool_before_file**: Before generating a specific file (inside Files array)
3. **tool_after_file**: After generating a specific file (inside Files array)
4. **tool_before_conclusion**: Before writing the Conclusion field

### Tool Call Mechanism:
- At any point, populate ONE tool field with: `{{"tool_name": "name", "query": "your query"}}`
- Set ALL OTHER fields to `null` (except content you're actively writing)
- After receiving tool results, set the used tool field to `null` and continue
- NEVER repeat content from previous responses - only fill new fields

**Avoid tool call for information provided in attached files.**

### Tool Field Selection Logic:

**Use tool_after_text when:**
- Need information that affects ALL files you'll generate
- Need to verify something before starting code generation
- Need current best practices that apply to entire solution

**Use tool_before_file when:**
- Need information specific to ONE file you're about to generate
- Need to check something before writing that particular file
- Rare - usually better to search before starting Files array

**Use tool_after_file when:**
- Generated a file and realized you need additional context for remaining files
- Need to verify something based on what you just wrote
- Rare - try to gather info upfront instead

**Use tool_before_conclusion when:**
- Need final verification before writing deployment/usage instructions
- Need to check something for the conclusion/guide
- Very rare - most info should be gathered earlier

### JSON Structure Example with Tool Flow:

**Initial Response (need info before files):**
```json
{{
  "Text": "I'll create a modern authentication system. Let me verify the latest security recommendations for 2025...",
  "tool_after_text": {{
    "tool_name": "search_web",
    "query": "OWASP authentication best practices 2025 JWT"
  }},
  "Files": null,
  "tool_before_conclusion": null,
  "Conclusion": null
}}
```

**After Tool Results (continue where you left off):**
```json
{{
  "Text": null,
  "tool_after_text": null,
  "Files": [
    {{
      "tool_before_file": null,
      "FileName": "auth.py",
      "FileVersion": "1",
      "FileCode": "# Code using 2025 best practices...",
      "FileText": "This implements OWASP 2025 recommendations...",
      "tool_after_file": null
    }}
  ],
  "tool_before_conclusion": null,
  "Conclusion": "Your authentication system follows current security standards..."
}}
```
```
**CRITICAL: email_tool is FULLY AGENTIC:**
- **ONE CALL FOR ENTIRE TASK**: Provide the COMPLETE multi-step task in a SINGLE tool call. NEVER split into multiple calls.
- **CONCISE QUERIES ONLY**: State WHAT you want done, not HOW. email_tool decides its own execution strategy.
- **NO EMAIL BODY COMPOSITION**: NEVER include full email content in the query. Just state the intent (e.g., "send reply about the meeting").
- **NO CONTEXT INJECTION**: Don't include conversation history in the query. email_tool can access chat history if needed.
- **TRUST THE AGENT**: email_tool will ask for missing info on its own.

## Important Guidelines
1.  **Prioritize Memory**: Always use long-term and short-term memory to inform responses.
2.  **Trust Recent Information**: Recent user messages take precedence over long-term memory.
3.  **File Content is Sacred**: If files are attached, that's your primary information source.
4.  **Be Context-Aware**: Don't mention the proprietary memory system.
5.  **Timestamps**: Use timestamps for time-related scenarios (convert to Pakistan time).
6.  **Markdown in Text Field**: Use proper markdown formatting in the Text field.
7.  **No Text Outside JSON**: NEVER write anything outside the JSON structure.

## Code Generation Guidelines
8. **Output Format**: Respond ONLY in valid JSON following CodeResponse schema.
9. **Code Quality**: Generate production-ready, well-commented, properly structured code.
10. **File Organization**: Create logical file structures when multiple files are needed.
11. **File Versioning**: New files get FileVersion "1", edited files increment version.
12. **No Assumptions**: Ask for clarification in Text field when requirements are ambiguous.
13. **Search Sparingly**: Default to reasoning and existing context over searching.

Current Date: {today}
User's Preferred Name: {user_name}
"""

# Enhanced Pydantic schemas with tool support
class ToolCall(BaseModel):
    tool_name: str = Field(description="Name of the tool to call (e.g., 'search_web')")
    query: str = Field(description="Query or input for the tool")

class CodeFile(BaseModel):
    tool_before_file: Optional[ToolCall] = Field(default=None, description="Tool call before this file")
    FileName: str = Field(description="The name of the file including extension")
    FileVersion: Optional[str] = Field(description="The version of the file, e.g., '1'")
    FileCode: str = Field(description="The complete content of the file")
    FileText: Optional[str] = Field(description="Any text required like explanation, note or anything for that file")
    tool_after_file: Optional[ToolCall] = Field(default=None, description="Tool call after this file")

class CodeResponse(BaseModel):
    Text: Optional[str] = Field(default=None, description="Optional explanation or description of the solution before files")
    tool_after_text: Optional[ToolCall] = Field(default=None, description="Tool call after text field")
    Files: Optional[List[CodeFile]] = Field(default=None, description="List of generated code files")
    tool_before_conclusion: Optional[ToolCall] = Field(default=None, description="Tool call before conclusion")
    Conclusion: Optional[str] = Field(default=None, description="Any text like explanation, description, conclusion, a guide, or anything else needed after project files")

# Continuation prompt template
CONTINUATION_PROMPT_TEMPLATE = """[CONTINUATION CONTEXT]

ORIGINAL USER REQUEST:
"{original_query}"

YOUR RESPONSE SO FAR (do NOT repeat this):
\"\"\"
{partial_response}
\"\"\"

TOOL CALL YOU JUST MADE:
{tool_call_json}

TOOL RESULTS:
{tool_result_json}

---
CRITICAL INSTRUCTIONS:
1. Review the ORIGINAL USER REQUEST above - that is your complete task
2. You have already written the text in "YOUR RESPONSE SO FAR" - do NOT repeat it
3. Use the tool results above to continue your response
4. If the original request has multiple parts/steps, make sure to address ALL of them
5. You can call additional tools if needed to fully complete the original request
6. To use a next tool just end your text on tool calling JSON as {{"tool_call": "search_web", "query": "..."}}. CRITICAL: Your tool calling JSON Must be exactly formatted otherwise it cause huge breakdown.
7. Continue writing naturally from where you stopped until the ENTIRE original request is satisfied.

Continue now:"""

# Code mode continuation prompt
CODE_CONTINUATION_PROMPT_TEMPLATE = """[CONTINUATION CONTEXT]

ORIGINAL USER REQUEST:
"{original_query}"

YOUR JSON RESPONSE SO FAR:
{partial_json}

TOOL CALL YOU MADE (from field: {tool_field_name}):
{tool_call_json}

TOOL RESULTS:
{tool_result_json}

---
**CRITICAL CONTINUATION INSTRUCTIONS:**

1. Review ORIGINAL REQUEST - is it now fully answerable?
2. Set "{tool_field_name}" to null (you already used it)
3. Set ALL previously filled fields to null (Text, Files, etc. - do NOT repeat)
4. Use tool results to continue populating remaining fields

**Before making another tool call, verify:**
- Is additional information ESSENTIAL (not just "nice to have")?
- Is it time-sensitive/post-cutoff (not existing knowledge)?
- Is it NOT in the attached files (if any)?
- If NO to any ΓåÆ Complete response without more tools

**To make another tool call (rare):**
Choose appropriate field based on where you are in the response:
- Still need info before files? ΓåÆ `tool_after_text`
- Need info for specific file? ΓåÆ `tool_before_file` or `tool_after_file`
- Need info for conclusion? ΓåÆ `tool_before_conclusion`

**Default: Complete the response with existing context and tool results.**

Continue your JSON response:"""

# [Previous utility functions remain the same: extract_text_from_pdf, extract_text_from_docx,
#  extract_text_from_xlsx, extract_file_content, format_file_size, create_stitched_prompt,
#  get_tokenizer_for_model, count_tokens, count_message_tokens, get_user_chat_settings,
#  validate_reason_parameter - keeping them exactly as they were]

def extract_file_content_from_bytes(file_bytes, mime_type):
    """Extract content from file bytes."""
    from routes.file_routes import extract_file_content_from_bytes as extract_func
    return extract_func(file_bytes, mime_type)

def extract_text_from_pdf(file_path):
    """Extract text from PDF with error handling."""
    try:
        import pypdf
        with open(file_path, 'rb') as f:
            pdf_reader = pypdf.PdfReader(f)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text()
            return text if text.strip() else "[PDF content could not be extracted]"
    except Exception as e:
        logging.warning(f"PDF extraction failed: {e}")
        return f"[PDF content extraction error: {str(e)}]"

def extract_text_from_docx(file_path):
    """Extract text from DOCX with error handling."""
    try:
        import docx
        doc = docx.Document(file_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text if text.strip() else "[DOCX content could not be extracted]"
    except Exception as e:
        logging.warning(f"DOCX extraction failed: {e}")
        return f"[DOCX content extraction error: {str(e)}]"

def extract_text_from_xlsx(file_path):
    """Extract text from XLSX with error handling."""
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file_path)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = ",".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
            text += "\n"
        return text if text.strip() else "[XLSX content could not be extracted]"
    except Exception as e:
        logging.warning(f"XLSX extraction failed: {e}")
        return f"[XLSX content extraction error: {str(e)}]"

def extract_file_content(file_path, mime_type):
    """Extract content from file based on type."""
    try:
        if mime_type == 'application/pdf':
            return extract_text_from_pdf(file_path)
        elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
            return extract_text_from_docx(file_path)
        elif mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            return extract_text_from_xlsx(file_path)
        else:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
    except UnicodeDecodeError:
        for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    return f.read()
            except:
                continue
        return "[Binary file - content not readable]"
    except Exception as e:
        logging.error(f"File content extraction error: {e}")
        return f"[Error reading file: {str(e)}]"

def format_file_size(size_bytes):
    """Format file size in human-readable format."""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} MB"

def create_stitched_prompt(user_text, file_data_list):
    """Create a stitched prompt with files content."""
    if not file_data_list:
        return f"[USER MESSAGE]\n{user_text}\n[ATTACHED FILES: 0]"

    stitched = f"[USER MESSAGE]\n{user_text}\n[ATTACHED FILES: {len(file_data_list)}]\n"

    for idx, file_data in enumerate(file_data_list, 1):
        file_size_str = format_file_size(file_data['size'])
        stitched += f"ΓöÇΓöÇΓöÇ FILE {idx}: {file_data['original_name']} ({file_data['mime_type']}, {file_size_str}) ΓöÇΓöÇΓöÇ\n"
        stitched += f"{file_data['content']}\n"
        stitched += f"ΓöÇΓöÇΓöÇ END FILE {idx} ΓöÇΓöÇΓöÇ\n"

    return stitched

def get_tokenizer_for_model(model_name):
    """Get appropriate tokenizer for the model."""
    try:
        model_tokenizer_map = {
            "meta-llama/Llama-3.3-70B-Instruct-Turbo": "cl100k_base",
            "meta-llama/Llama-3.3-70B-Instruct-Turbo-Free": "cl100k_base",
            "meta-llama/Meta-Llama-3.1-8B-Instruct-Turbo": "cl100k_base",
            "meta-llama/Meta-Llama-3.1-70B-Instruct-Turbo": "cl100k_base",
            "Qwen/Qwen3-235B-A22B-fp8-tput": "cl100k_base",
            "Qwen/Qwen2.5-VL-72B-Instruct": "cl100k_base",
            "Qwen/Qwen2.5-72B-Instruct-Turbo": "cl100k_base",
            "Qwen/Qwen3-Coder-480B-A35B-Instruct-FP8": "cl100k_base",
            "default": "cl100k_base"
        }
        encoding_name = model_tokenizer_map.get(model_name, "cl100k_base")
        return tiktoken.get_encoding(encoding_name)
    except Exception as e:
        logging.warning(f"Failed to get tokenizer for {model_name}: {e}. Using default.")
        return tiktoken.get_encoding("cl100k_base")

def count_tokens(text, model_name):
    """Count tokens in text using appropriate tokenizer for the model."""
    try:
        if not text or not isinstance(text, str):
            return 0
        tokenizer = get_tokenizer_for_model(model_name)
        return len(tokenizer.encode(text))
    except Exception as e:
        logging.error(f"Token counting failed for model {model_name}: {e}")
        return max(1, len(text) // 4)

def count_message_tokens(messages, model_name):
    """Count tokens in a list of messages."""
    try:
        total_tokens = 0
        for message in messages:
            total_tokens += 4
            content = message.get('content', '')
            if isinstance(content, str):
                total_tokens += count_tokens(content, model_name)
            elif isinstance(content, list):
                for item in content:
                    if item.get('type') == 'text':
                        total_tokens += count_tokens(item.get('text', ''), model_name)
                    elif item.get('type') == 'image_url':
                        total_tokens += 765
        return total_tokens
    except Exception as e:
        logging.error(f"Message token counting failed: {e}")
        total_text = ' '.join([str(msg.get('content', '')) for msg in messages])
        return max(10, len(total_text) // 4)

def get_or_create_anonymous_user(session_id):
    """
    Get or create an anonymous user record in the database for unauthenticated sessions.
    Returns the anonymous user's ID or None if rate limit is exceeded.
    """
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        # Check if we already have an anonymous user for this session_id
        # We'll create a special email pattern for anonymous users
        anonymous_email = f"anonymous_{session_id}@anonymous.local"
        
        # Try to get existing anonymous user
        cursor.execute(
            "SELECT id FROM users WHERE email = %s",
            (anonymous_email,)
        )
        result = cursor.fetchone()
        
        if result:
            return result['id']
        
        # Create a new anonymous user
        cursor.execute(
            "INSERT INTO users (email, username, password) VALUES (%s, %s, %s) RETURNING id",
            (anonymous_email, f"Anonymous_{session_id[:8]}", None)
        )
        new_user = cursor.fetchone()
        
        if new_user:
            return new_user['id']
        else:
            return None
            
    except Exception as e:
        logging.error(f"Error creating anonymous user: {e}", exc_info=True)
        return None
    finally:
        return_db_connection(conn)

def check_user_token_limit(user_id):
    """
    Check if user has exceeded their free token allotment.
    Returns (has_exceeded, used_tokens, limit) tuple.
    """
    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Get total tokens used by this user
        cursor.execute(
            """SELECT SUM(input_tokens + output_tokens) as total_tokens
            FROM token_usage
            WHERE user_id = %s""",
            (user_id,)
        )
        result = cursor.fetchone()
        used_tokens = int(result['total_tokens'] or 0)

        # Get token limit from config
        token_limit = current_app.config['FREE_TOKEN_ALLOTMENT']

        has_exceeded = used_tokens >= token_limit

        return (has_exceeded, used_tokens, token_limit)

    except Exception as e:
        logging.error(f"Error checking token limit: {e}", exc_info=True)
        return (False, 0, 0)  # Allow on error to avoid blocking users
    finally:
        return_db_connection(conn)

def get_user_chat_settings(user_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT temperature, top_p, system_prompt, what_we_call_you, together_api_key FROM user_settings WHERE user_id = %s", 
            (user_id,)
        )
        settings = cursor.fetchone()
        if settings:
            return {
                "temperature": settings['temperature'] if settings['temperature'] is not None else 0.7,
                "top_p": settings['top_p'] if settings['top_p'] is not None else 1.0,
                "system_prompt": settings['system_prompt'] or "You are a helpful assistant.",
                "what_we_call_you": settings['what_we_call_you'] or "User",
                "together_api_key": (decrypt_key(settings['together_api_key']) if settings['together_api_key'] else None)
            }
    finally:
        return_db_connection(conn)
    return {"temperature": 0.7, "top_p": 1.0, "system_prompt": "You are a helpful assistant.", "what_we_call_you": "User", "together_api_key": None}

def validate_reason_parameter(reason):
    """Validate and normalize the reason parameter."""
    if reason is None:
        return "default"
    if isinstance(reason, bool):
        return "reason" if reason else "default"
    if isinstance(reason, str):
        reason = reason.lower().strip()
        if reason in ["code", "reason", "default"]:
            return reason
        else:
            logging.warning(f"Invalid reason parameter: {reason}. Defaulting to 'default'")
            return "default"
    logging.warning(f"Unexpected reason parameter type: {type(reason)}. Defaulting to 'default'")
    return "default"

def detect_tool_call_in_default(text: str) -> Optional[Dict[str, Any]]:
    """
    Detect tool call JSON in default mode response.
    More robust detection focusing on exact field names.
    Returns tool call dict if found, None otherwise.
    """
    try:
        text = text.strip()

        # Pattern: exact match for tool call structure
        # Looks for: {"tool_call": "search_web", "query": "..."}
        pattern = r'\{\s*"tool_call"\s*:\s*"([^"]+)"\s*,\s*"query"\s*:\s*"([^"]+)"\s*\}'

        # Search from the end (last occurrence)
        matches = list(re.finditer(pattern, text))

        if matches:
            last_match = matches[-1]
            tool_name = last_match.group(1)
            query = last_match.group(2)

            # Verify it's actually at the end (allow trailing whitespace/punctuation)
            remaining_text = text[last_match.end():].strip()
            if len(remaining_text) <= 2:  # Allow for trailing period or similar
                return {
                    "tool_call": tool_name,
                    "query": query
                }

        # Fallback: try JSON parsing from the end
        if text.endswith('}'):
            brace_count = 0
            json_start = -1
            for i in range(len(text) - 1, -1, -1):
                if text[i] == '}':
                    brace_count += 1
                elif text[i] == '{':
                    brace_count -= 1
                    if brace_count == 0:
                        json_start = i
                        break

            if json_start != -1:
                potential_json = text[json_start:]
                try:
                    parsed = json.loads(potential_json)
                    if 'tool_call' in parsed and 'query' in parsed:
                        # Verify these are the only keys or close to it
                        if len(parsed) == 2:
                            return parsed
                except json.JSONDecodeError:
                    pass
    except Exception as e:
        logging.debug(f"Tool call detection error: {e}")

    return None


def extract_text_before_tool_call(text: str) -> str:
    """
    Extract text before tool call JSON, handling edge cases.
    """
    try:
        # Use the same pattern as detection
        pattern = r'\{\s*"tool_call"\s*:\s*"[^"]+"\s*,\s*"query"\s*:\s*"[^"]+"\s*\}'

        matches = list(re.finditer(pattern, text))
        if matches:
            last_match = matches[-1]
            return text[:last_match.start()].strip()

        # Fallback
        json_start = text.rfind('{"tool_call"')
        if json_start != -1:
            return text[:json_start].strip()

        return text
    except Exception as e:
        logging.warning(f"Tool call extraction failed: {e}")
        return text



def detect_tool_call_in_code(json_obj: Dict[str, Any]) -> Optional[tuple]:
    """
    Detect tool call in code mode JSON.
    Returns (field_name, tool_call_dict) if found, None otherwise.
    """
    tool_fields = ['tool_after_text', 'tool_before_conclusion']

    for field in tool_fields:
        if field in json_obj and json_obj[field] is not None:
            return (field, json_obj[field])

    # Check in Files array
    files = json_obj.get('Files', [])
    if files:
        for idx, file_obj in enumerate(files):
            if isinstance(file_obj, dict):
                for field in ['tool_before_file', 'tool_after_file']:
                    if field in file_obj and file_obj[field] is not None:
                        return (f"Files[{idx}].{field}", file_obj[field])

    return None

def merge_json_responses(responses: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Merge multiple JSON responses, keeping last non-null value for each field.
    """
    merged = {}

    for response in responses:
        for key, value in response.items():
            if value is not None:
                if key == 'Files' and isinstance(value, list):
                    # Special handling for Files array
                    if 'Files' not in merged:
                        merged['Files'] = []

                    for file_obj in value:
                        if file_obj is not None:
                            # Remove tool fields before merging
                            clean_file = {k: v for k, v in file_obj.items() if v is not None}
                            if clean_file:
                                merged['Files'].append(clean_file)
                else:
                    # Keep ALL fields including tool fields for history
                    merged[key] = value

    return merged


def extract_essential_search_results(tavily_response: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract only the essential information from Tavily response for LLM context.
    Includes citation indices for source attribution.
    """
    essential = {}

    # Include the query for reference
    if 'query' in tavily_response:
        essential['query'] = tavily_response['query']

    # Include the answer if available (this is the most important!)
    if tavily_response.get('answer'):
        essential['answer'] = tavily_response['answer']

    # Include top 3 results with citation indices
    if 'results' in tavily_response and tavily_response['results']:
        essential['results'] = []
        for idx, result in enumerate(tavily_response['results'][:3], start=1):  # 1-based indexing
            essential['results'].append({
                'index': idx,  # ΓåÉ NEW: Citation index
                'title': result.get('title', ''),
                'url': result.get('url', ''),
                'content': result.get('content', '')[:1000]  # Limit to 1000 chars
            })

    return essential

def extract_urls_from_tavily_response(tavily_response: Dict[str, Any]) -> List[Dict[str, str]]:
    """
    Extract URLs from Tavily response with citation indices.
    Returns list of {index, url, title} objects.
    """
    urls = []
    results = tavily_response.get('results', [])
    
    for idx, result in enumerate(results, start=1):  # 1-based to match LLM citations
        if 'url' in result:
            urls.append({
                'index': idx,  # ΓåÉ Citation index
                'url': result['url'],
                'title': result.get('title', 'Untitled')
            })
    
    return urls

def store_search_web_urls(user_id, session_id, chat_history_id, search_calls):
    """
    Store search_web URLs in database.
    
    Args:
        user_id: User ID
        session_id: Session number
        chat_history_id: Chat history record ID
        search_calls: List of {query, urls, timestamp} dicts
    """
    if not search_calls:
        return
    
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        for idx, call in enumerate(search_calls):
            cursor.execute(
                """INSERT INTO search_web_logs
                   (user_id, session_number, chat_history_id, call_sequence, query, urls_json, timestamp)
                   VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                (user_id, int(session_id), chat_history_id, idx, 
                 call['query'], json.dumps(call['urls']), call['timestamp'])
            )
        conn.commit()
        logging.info(f"Stored {len(search_calls)} search_web URL logs for chat_history_id {chat_history_id}")
    except Exception as e:
        conn.rollback()
        logging.error(f"Failed to store search_web URLs: {e}", exc_info=True)
    finally:
        return_db_connection(conn)


def store_email_tool_data(user_id, session_id, chat_history_id, email_tool_data):
    """
    Store email_tool execution data in database for history UI reconstruction.
    
    Args:
        user_id: User ID
        session_id: Session number
        chat_history_id: Chat history record ID
        email_tool_data: Dict with {query, success, total_iterations, summary, iterations, timestamp}
    """
    if not email_tool_data:
        return
    
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            """INSERT INTO email_tool_logs
               (user_id, session_number, chat_history_id, query, success, total_iterations, summary, iterations_json, timestamp)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (user_id, int(session_id), chat_history_id, 
             email_tool_data.get('query', ''),
             email_tool_data.get('success', True),
             email_tool_data.get('total_iterations', 0),
             email_tool_data.get('summary', ''),
             json.dumps(email_tool_data.get('iterations', [])),
             email_tool_data.get('timestamp', datetime.now(timezone.utc).isoformat()))
        )
        conn.commit()
        logging.info(f"Stored email_tool data for chat_history_id {chat_history_id}")
    except Exception as e:
        conn.rollback()
        logging.error(f"Failed to store email_tool data: {e}", exc_info=True)
    finally:
        return_db_connection(conn)


@chat_bp.route('/chat', methods=['POST'])
@optional_token_required
def chat(current_user):
    data = request.json or {}
    session_id = data.get('session_id')
    query = data.get('query', '').strip()
    
    # Extract client context for timezone-aware tool operations
    # Frontend should send: client_datetime (ISO string), client_timezone (IANA timezone)
    client_context = {
        'local_datetime': data.get('client_datetime'),
        'timezone': data.get('client_timezone')
    }

    if not session_id or not query:
        return jsonify({"error": "session_id and query are required"}), 400

    is_vision_request = False
    file_data_list = []
    image_url = None

    if current_user:
        user_id = current_user['id']
        cache_key = f"{user_id}-{session_id}"

        # Check for staged files
        if hasattr(current_app, 'file_cache') and cache_key in current_app.file_cache:
            file_ids = current_app.file_cache.pop(cache_key)
            logging.info(f"Found {len(file_ids)} staged files in cache for {cache_key}")

            conn = get_db_connection()
            try:
                cursor = conn.cursor()
                placeholders = ','.join(['%s'] * len(file_ids))
                cursor.execute(
                    f"""SELECT id, b2_key, original_name, size, mime_type, is_image
                       FROM uploaded_files
                       WHERE id IN ({placeholders})""",
                    file_ids
                )
                files = cursor.fetchall()
                # Helper function to download from B2
                def download_from_b2(b2_key):
                    """Download file from B2 using presigned URL."""
                    from routes.file_routes import generate_presigned_url
                    url = generate_presigned_url(b2_key, expiration=600)  # 10 minutes
                    if not url:
                        raise Exception(f"Failed to generate presigned URL for {b2_key}")
    
                    response = requests.get(url, timeout=30)
                    response.raise_for_status()
                    return response.content

                for file_record in files:
                    b2_key = file_record['b2_key']
    
                    try:
                        file_bytes = download_from_b2(b2_key)
                    except Exception as e:
                        logging.error(f"Failed to download file from B2: {b2_key}, error: {e}", exc_info=True)
                        file_data_list.append({
                            'id': file_record['id'],
                            'b2_key': b2_key,
                            'original_name': file_record['original_name'],
                            'size': file_record['size'],
                            'mime_type': file_record['mime_type'],
                            'content': f"[Error: Failed to load file {file_record['original_name']}]"
                        })
                        continue

                    if file_record['is_image']:
                        import base64
                        encoded = base64.b64encode(file_bytes).decode('utf-8')
                        image_url = f"data:{file_record['mime_type']};base64,{encoded}"
                        is_vision_request = True
                    else:
                        content = extract_file_content_from_bytes(file_bytes, file_record['mime_type'])
                        logging.info(f"Extracted content from {file_record['original_name']}: {len(content)} characters")
                        file_data_list.append({
                            'id': file_record['id'],
                            'b2_key': b2_key,
                            'original_name': file_record['original_name'],
                            'size': file_record['size'],
                            'mime_type': file_record['mime_type'],
                            'content': content
                        })    
            finally:
                return_db_connection(conn)

        reason = validate_reason_parameter(data.get('reason'))
        chat_settings = get_user_chat_settings(user_id)
        api_key = chat_settings.get('together_api_key') or current_app.config['TOGETHER_API_KEY']

        # Check if user has their own API key - if not, check token limits
        if not chat_settings.get('together_api_key'):
            has_exceeded, used_tokens, token_limit = check_user_token_limit(user_id)
            if has_exceeded:
                return jsonify({
                    "error": "What?? It seems you've used all your tokens!! Don't you have shame? 😱 It's time to use your own API key; navigate to together.ai, get your API key, and paste that key in Settings → Together API Key!",
                    "token_limit_exceeded": True,
                    "used_tokens": used_tokens,
                    "token_limit": token_limit
                }), 429
    else:
        # Unauthenticated users get a humorous rejection message
        return jsonify({
            "error": "Are you kidding me? How can you be so frank if I don't know you? Please login first! 😤",
            "login_required": True
        }), 401

    memory = TokenAwareMemoryManager(user_id, session_id)
    client = Together(api_key=api_key)
    original_prompt = query

    # Determine model and prepare messages
    if is_vision_request:
        model_name = "Qwen/Qwen2.5-VL-72B-Instruct"
        messages = [{
            "role": "user",
            "content": [
                {"type": "text", "text": query},
                {"type": "image_url", "image_url": {"url": image_url}}
            ]
        }]
    else:
        stitched_prompt = create_stitched_prompt(query, file_data_list)
        context_messages = memory.get_context()
        context_messages.append({"role": "user", "content": stitched_prompt})

        if reason == "code":
            model_name = current_app.config['CODE_LLM']
            final_system_prompt = CODE_SYSTEM_PROMPT_TEMPLATE.format(today=current_date(), user_name=chat_settings['what_we_call_you'])
        elif reason == "reason":
            model_name = current_app.config['REASON_LLM']
            final_system_prompt = BASE_SYSTEM_PROMPT.format(
                today=current_date(),
                user_name=chat_settings['what_we_call_you'],
                user_persona=chat_settings['system_prompt']
            )
        else:
            model_name = current_app.config['DEFAULT_LLM']
            final_system_prompt = BASE_SYSTEM_PROMPT.format(
                today=current_date(),
                user_name=chat_settings['what_we_call_you'],
                user_persona=chat_settings['system_prompt']
            )

        messages = [{"role": "system", "content": final_system_prompt}] + context_messages

    def generate_and_update_memory():
        generation_completed_normally = False
        max_tool_calls = current_app.config.get('MAX_TOOL_CALLS_PER_INTERACTION', 5)
        tool_call_count = 0

        # Track responses for code mode merging
        code_mode_responses = []

        # For default mode, track cumulative response
        default_mode_full_response = ""

        search_web_calls = []  # Track search_web executions: [{query, urls, timestamp}]
        email_tool_data = None  # Track email_tool execution: {query, success, total_iterations, summary, iterations, timestamp}

        try:
            # Main tool loop
            while tool_call_count < max_tool_calls:
                chunks = []
                current_messages = messages.copy()

                logging.info(f"Tool loop iteration {tool_call_count + 1}, mode: {reason}")

                # Prepare request parameters
                request_params = {
                    "model": model_name,
                    "messages": current_messages,
                    "temperature": chat_settings['temperature'],
                    "top_p": chat_settings['top_p'],
                    "max_tokens": 10000,
                    "stream": True
                }

                if reason == "code":
                    request_params["response_format"] = {
                        "type": "json_schema",
                        "schema": CodeResponse.model_json_schema(),
                    }

                # Stream response
                stream = client.chat.completions.create(**request_params)
                import sys
                for token_obj in stream:
                    if token_obj.choices:
                        delta = token_obj.choices[0].delta.content or ''
                        chunks.append(delta)
                        data = f"data: {json.dumps({'token': delta, 'mode': reason})}\n\n".encode()
                        yield data
                        # force flush to prevent buffering...
                        try: 
                            sys.stdout.flush()
                        except: 
                            pass
                partial_response = ''.join(chunks).strip()

                if not partial_response:
                    logging.warning("Empty response received, breaking tool loop")
                    break

                # Process based on mode
                if reason == "code":
                    # Code mode: detect tool in JSON
                    try:
                        json_response = json.loads(partial_response)
                        code_mode_responses.append(json_response)

                        tool_detection = detect_tool_call_in_code(json_response)

                        if tool_detection:
                            field_name, tool_call_data = tool_detection
                            tool_name = tool_call_data.get('tool_name')
                            tool_query = tool_call_data.get('query')

                            logging.info(f"Tool call detected in code mode: {tool_name} from field {field_name}")

                            # Send tool call event
                            yield f"data: {json.dumps({'event': 'tool_call', 'tool_name': tool_name, 'mode': reason})}\n\n".encode()

                            # Execute tool
                            loop = asyncio.new_event_loop()
                            asyncio.set_event_loop(loop)
                            tool_result = loop.run_until_complete(
                                execute_tool(tool_name, {'query': tool_query}, user_id=user_id, session_id=str(session_id), socketio_instance=current_app.socketio if hasattr(current_app, 'socketio') else None, client_context=client_context)
                            )
                            loop.close()
                            # Track search_web URLs
                            if tool_name == 'search_web' and tool_result.get('success'):
                                urls = extract_urls_from_tavily_response(tool_result['result'])
                                search_web_calls.append({
                                    'query': tool_query,
                                    'urls': urls,
                                    'timestamp': datetime.now(timezone.utc).isoformat()
                                })
                                logging.info(f"Captured {len(urls)} URLs from search_web call #{len(search_web_calls)}")
                                # After capturing URLs, add to database cache for cross-worker access
                                if search_web_calls:
                                    try:
                                        conn = get_db_connection()
                                        cursor = conn.cursor()
                                        cursor.execute(
                                            """INSERT INTO search_web_realtime_cache (user_id, session_number, calls_json, updated_at)
                                                VALUES (%s, %s, %s, NOW())
                                                ON CONFLICT (user_id, session_number) 
                                                DO UPDATE SET calls_json = EXCLUDED.calls_json, updated_at = NOW()""",
                                            (user_id, int(session_id), json.dumps(search_web_calls))
                                        )   
                                        conn.commit()
                                        logging.info(f"Updated realtime cache for session {session_id} with {len(search_web_calls)} calls")
                                        return_db_connection(conn)
                                    except Exception as e:
                                        logging.error(f"Failed to update realtime cache: {e}", exc_info=True)
                                        if conn:
                                            return_db_connection(conn)

                            # Track email_tool data for history persistence
                            if tool_name == 'email_tool' and tool_result.get('success'):
                                result_data = tool_result.get('result', {})
                                email_tool_data = {
                                    'query': tool_query,
                                    'success': result_data.get('success', True),
                                    'total_iterations': result_data.get('total_iterations', 0),
                                    'summary': result_data.get('summary', ''),
                                    'iterations': result_data.get('iterations', []),
                                    'timestamp': datetime.now(timezone.utc).isoformat()
                                }
                                logging.info(f"Captured email_tool data with {email_tool_data['total_iterations']} iterations")

                            if not tool_result.get('success'):
                                logging.error(f"Tool execution failed: {tool_result.get('error')}")
                                # Continue without tool result
                                break

                            tool_call_count += 1

                            # Extract only essential search results
                            essential_results = tool_result.get('result', tool_result) if tool_name == 'email_tool' else extract_essential_search_results(tool_result['result'])
                            # Prepare continuation prompt
                            continuation_prompt = CODE_CONTINUATION_PROMPT_TEMPLATE.format(
                                original_query=original_prompt,
                                partial_json=json.dumps(json_response, indent=2),
                                tool_field_name=field_name,
                                tool_call_json=json.dumps(tool_call_data, indent=2),
                                tool_result_json=json.dumps(essential_results, indent=2)
                            )

                            # Add continuation to messages
                            messages.append({"role": "assistant", "content": partial_response})
                            messages.append({"role": "user", "content": continuation_prompt})

                            # Continue loop for next iteration
                            continue
                        else:
                            # No more tool calls, response complete
                            break

                    except json.JSONDecodeError as e:
                        logging.error(f"Invalid JSON in code mode: {e}")
                        yield f"data: {json.dumps({'error': 'Invalid JSON generated', 'mode': reason})}\n\n".encode()
                        break

                else:
                    # Default/Reason mode: detect tool call at end
                    default_mode_full_response += partial_response

                    tool_call_data = detect_tool_call_in_default(partial_response)

                    if tool_call_data:
                        logging.info(f"=== TOOL CALL DETECTION START ===")
                        logging.info(f"Raw tool_call_data type: {type(tool_call_data)}")
                        logging.info(f"Raw tool_call_data: {tool_call_data}")
                        logging.info(f"tool_call_data keys: {list(tool_call_data.keys()) if isinstance(tool_call_data, dict) else 'NOT A DICT'}")

                        tool_name = tool_call_data.get('tool_call')
                        tool_query = tool_call_data.get('query')

                        logging.info(f"Extracted tool_name: {tool_name}")
                        logging.info(f"Extracted tool_query: {tool_query}")
                        logging.info(f"=== TOOL CALL DETECTION END ===")

                        # Remove tool call JSON from partial response
                        text_before_tool = extract_text_before_tool_call(partial_response)

                        # Send tool call event
                        yield f"data: {json.dumps({'event': 'tool_call', 'tool_name': tool_name, 'mode': reason})}\n\n".encode()

                        # Execute tool
                        # Execute tool with detailed error handling
                        try:
                            logging.info(f"=== TOOL EXECUTION START ===")
                            logging.info(f"Calling execute_tool with: tool_name={tool_name}, query={tool_query}")

                            loop = asyncio.new_event_loop()
                            asyncio.set_event_loop(loop)
                            tool_result = loop.run_until_complete(
                                execute_tool(tool_name, {'query': tool_query}, user_id=user_id, session_id=str(session_id), socketio_instance=current_app.socketio if hasattr(current_app, 'socketio') else None, client_context=client_context)
                            )
                            loop.close()
                            # Track search_web URLs
                            if tool_name == 'search_web' and tool_result.get('success'):
                                urls = extract_urls_from_tavily_response(tool_result['result'])
                                search_web_calls.append({
                                    'query': tool_query,
                                    'urls': urls,
                                    'timestamp': datetime.now(timezone.utc).isoformat()
                                })
                                logging.info(f"Captured {len(urls)} URLs from search_web call #{len(search_web_calls)}")
                                # After capturing URLs, add to database cache for cross-worker access
                                if search_web_calls:
                                    try:
                                        conn = get_db_connection()
                                        cursor = conn.cursor()
                                        cursor.execute(
                                            """INSERT INTO search_web_realtime_cache (user_id, session_number, calls_json, updated_at)
                                                VALUES (%s, %s, %s, NOW())
                                                ON CONFLICT (user_id, session_number) 
                                                DO UPDATE SET calls_json = EXCLUDED.calls_json, updated_at = NOW()""",
                                            (user_id, int(session_id), json.dumps(search_web_calls))
                                        )
                                        conn.commit()
                                        logging.info(f"Updated realtime cache for session {session_id} with {len(search_web_calls)} calls")
                                        return_db_connection(conn)
                                    except Exception as e:
                                        logging.error(f"Failed to update realtime cache: {e}", exc_info=True)
                                        if conn:
                                            return_db_connection(conn)

                            # Track email_tool data for history persistence
                            if tool_name == 'email_tool' and tool_result.get('success'):
                                result_data = tool_result.get('result', {})
                                email_tool_data = {
                                    'query': tool_query,
                                    'success': result_data.get('success', True),
                                    'total_iterations': result_data.get('total_iterations', 0),
                                    'summary': result_data.get('summary', ''),
                                    'iterations': result_data.get('iterations', []),
                                    'timestamp': datetime.now(timezone.utc).isoformat()
                                }
                                logging.info(f"Captured email_tool data with {email_tool_data['total_iterations']} iterations")

                            logging.info(f"Tool result type: {type(tool_result)}")
                            logging.info(f"Tool result keys: {list(tool_result.keys()) if isinstance(tool_result, dict) else 'NOT A DICT'}")
                            logging.info(f"Tool result: {tool_result}")
                            logging.info(f"=== TOOL EXECUTION END ===")

                            if not tool_result.get('success'):
                                logging.error(f"Tool execution failed: {tool_result.get('error')}")
                                error_msg = f"\n\n*[Tool execution failed: {tool_result.get('error', 'Unknown error')}]*"
                                yield f"data: {json.dumps({'token': error_msg, 'mode': reason})}\n\n".encode()
                                break

                        except Exception as tool_exec_error:
                            logging.error(f"!!! TOOL EXECUTION CRASHED !!!", exc_info=True)
                            logging.error(f"Error type: {type(tool_exec_error)}")
                            logging.error(f"Error message: {str(tool_exec_error)}")
                            error_msg = f"\n\n*[Tool execution crashed: {str(tool_exec_error)}]*"
                            yield f"data: {json.dumps({'token': error_msg, 'mode': reason})}\n\n".encode()
                            break

                        tool_call_count += 1

                        # Prepare continuation prompt
                        # Prepare continuation prompt with error handling
                        try:
                            logging.info(f"=== CONTINUATION PROMPT CREATION START ===")
                            logging.info(f"original_prompt: {original_prompt[:100]}...")
                            logging.info(f"text_before_tool length: {len(text_before_tool)}")
                            logging.info(f"tool_call_data: {tool_call_data}")

                            # Check tool_result structure
                            if 'result' in tool_result:
                                tool_result_data = tool_result['result']
                                logging.info(f"Using tool_result['result']")
                            else:
                                logging.warning(f"!!! tool_result missing 'result' key, using full tool_result")
                                logging.warning(f"Available keys: {list(tool_result.keys())}")
                                tool_result_data = tool_result

                            # Extract only essential search results
                            essential_results = tool_result.get('result', tool_result) if tool_name == 'email_tool' else extract_essential_search_results(tool_result['result'])

                            continuation_prompt = CONTINUATION_PROMPT_TEMPLATE.format(
                                original_query=original_prompt,
                                partial_response=text_before_tool,
                                tool_call_json=json.dumps(tool_call_data, indent=2),
                                tool_result_json=json.dumps(essential_results, indent=2)  # ΓåÉ MUCH SMALLER!
                            )

                            logging.info(f"Essential results size: {len(json.dumps(essential_results))} chars (vs full: {len(json.dumps(tool_result['result']))} chars)")

                            logging.info(f"Continuation prompt created successfully, length: {len(continuation_prompt)}")
                            logging.info(f"=== CONTINUATION PROMPT CREATION END ===")

                        except KeyError as ke:
                            logging.error(f"!!! KEYERROR IN CONTINUATION PROMPT !!!", exc_info=True)
                            logging.error(f"Missing key: {ke}")
                            logging.error(f"tool_result structure: {tool_result}")
                            raise
                        except Exception as cont_error:
                            logging.error(f"!!! CONTINUATION PROMPT CREATION CRASHED !!!", exc_info=True)
                            logging.error(f"Error: {cont_error}")
                            raise

                        # Update messages for next iteration
                        messages.append({"role": "assistant", "content": text_before_tool})
                        messages.append({"role": "user", "content": continuation_prompt})

                        # Continue loop
                        continue
                    else:
                        # No tool call, response complete
                        break

            # Response generation complete
            generation_completed_normally = True

        except GeneratorExit:
            logging.warning(f"Client disconnected, generation for session {session_id} was interrupted.")
        except Exception as e:
            logging.error(f"Streaming error: {e}", exc_info=True)
            error_response = {'error': 'Generation failed', 'details': str(e), 'mode': reason}
            yield f"data: {json.dumps(error_response)}\n\n".encode()

        finally:
            if generation_completed_normally:
                # Save to memory based on mode
                if reason == "code" and code_mode_responses:
                    # Merge all JSON responses
                    final_json = merge_json_responses(code_mode_responses)
                    final_response = json.dumps(final_json, indent=2)

                    memory_query = stitched_prompt if file_data_list else original_prompt
                    input_token_count = count_tokens(memory_query, model_name)
                    output_token_count = count_tokens(final_response, model_name)

                    memory.add_interaction(memory_query, final_response, input_token_count, output_token_count, original_prompt=original_prompt)

                    # Link files to chat
                    if current_user:
                        conn = get_db_connection()
                        try:
                            cursor = conn.cursor()
                            cursor.execute(
                                """SELECT id FROM chat_history
                                   WHERE user_id = %s AND session_number = %s
                                   ORDER BY id DESC LIMIT 1""",
                                (user_id, session_id)
                            )
                            result = cursor.fetchone()
                            if not result: 
                                logging.warning(f"No chat history found for user {user_id}, session {session_id}")
                                return_db_connection(conn)
                                return

                            last_chat_id = result['id']
                            if file_data_list:
                                for file_data in file_data_list:
                                    cursor.execute(
                                        "INSERT INTO chat_files (chat_history_id, file_id) VALUES (%s, %s)",
                                        (last_chat_id, file_data['id'])
                                    )
                            conn.commit()
                        finally:
                            return_db_connection(conn)

                            # Store search_web URLs
                        if search_web_calls:
                            store_search_web_urls(user_id, session_id, last_chat_id, search_web_calls)
                        
                        # Store email_tool data
                        if email_tool_data:
                            store_email_tool_data(user_id, session_id, last_chat_id, email_tool_data)

                    logging.info(f"Added code interaction with tool usage: {output_token_count} tokens")

                elif reason == "reason" and not is_vision_request:
                    cleaned_answer = THINK_TAG_REGEX.sub('', default_mode_full_response).strip()
                    memory_query = stitched_prompt if file_data_list else original_prompt
                    input_token_count = count_tokens(memory_query, model_name)
                    output_token_count = count_tokens(cleaned_answer, model_name)

                    memory.add_interaction(memory_query, cleaned_answer, input_token_count, output_token_count,
                                         full_response_for_history=default_mode_full_response,
                                         original_prompt=original_prompt)

                    # Link files
                    if current_user:
                        conn = get_db_connection()
                        try:
                            cursor = conn.cursor()
                            cursor.execute(
                                """SELECT id FROM chat_history
                                   WHERE user_id = %s AND session_number = %s
                                   ORDER BY id DESC LIMIT 1""",
                                (user_id, session_id)
                            ) 
                            result = cursor.fetchone()
                            if not result:  # ADD THIS CHECK
                                logging.warning(f"No chat history found for user {user_id}, session {session_id}")
                                return_db_connection(conn)
                                return

                            last_chat_id = result['id']
                            if file_data_list:
                                for file_data in file_data_list:
                                    cursor.execute(
                                        "INSERT INTO chat_files (chat_history_id, file_id) VALUES (%s, %s)",
                                        (last_chat_id, file_data['id'])
                                    )
                            conn.commit()
                        finally:
                            return_db_connection(conn)

                        # Store search_web URLs
                        if search_web_calls:
                            store_search_web_urls(user_id, session_id, last_chat_id, search_web_calls)
                        
                        # Store email_tool data
                        if email_tool_data:
                            store_email_tool_data(user_id, session_id, last_chat_id, email_tool_data)

                    logging.info(f"Added reasoning interaction with tool usage: {output_token_count} tokens")

                else:
                    # Default mode or vision
                    final_response = default_mode_full_response if default_mode_full_response else ''.join(chunks).strip()
                    memory_query = f"[Image Analysis] {original_prompt}" if is_vision_request else (stitched_prompt if file_data_list else original_prompt)

                    input_token_count = count_tokens(memory_query, model_name)
                    output_token_count = count_tokens(final_response, model_name)

                    memory.add_interaction(memory_query, final_response, input_token_count, output_token_count, original_prompt=original_prompt)

                    # Link files
                    if current_user:
                        conn = get_db_connection()
                        try:
                            cursor = conn.cursor()
                            cursor.execute(
                                """SELECT id FROM chat_history
                                   WHERE user_id = %s AND session_number = %s
                                   ORDER BY id DESC LIMIT 1""",
                                (user_id, session_id)
                            )
                            result = cursor.fetchone()
                            if not result: 
                                logging.warning(f"No chat history found for user {user_id}, session {session_id}")
                                return_db_connection(conn)
                                return

                            last_chat_id = result['id']
                            if file_data_list:
                                for file_data in file_data_list:
                                    cursor.execute(
                                        "INSERT INTO chat_files (chat_history_id, file_id) VALUES (%s, %s)",
                                        (last_chat_id, file_data['id'])
                                    )
                            conn.commit()
                        finally:
                            return_db_connection(conn)

                        # Store search_web URLs
                        if search_web_calls:
                            store_search_web_urls(user_id, session_id, last_chat_id, search_web_calls)
                        
                        # Store email_tool data
                        if email_tool_data:
                            store_email_tool_data(user_id, session_id, last_chat_id, email_tool_data)

                    logging.info(f"Added default interaction with tool usage: {output_token_count} tokens")

                memory.save_to_db()

                # Send memory stats and completion
                memory_stats = memory.get_memory_stats()
                memory_stats['mode'] = reason
                yield f"data: {json.dumps({'memory_stats': memory_stats})}\n\n".encode()
                yield f"data: {json.dumps({'status': 'done', 'mode': reason})}\n\n".encode()
            else:
                logging.info(f"Generation for session {session_id} did not complete normally.")
            # Clear search_web realtime cache from database
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "DELETE FROM search_web_realtime_cache WHERE user_id = %s AND session_number = %s",
                    (user_id, int(session_id))
                )
                conn.commit()
                logging.info(f"Cleared realtime cache for session {session_id}")
                return_db_connection(conn)
            except Exception as e:
                logging.error(f"Failed to clear realtime cache: {e}", exc_info=True)
                if conn:
                    return_db_connection(conn)
            yield b"event: end-of-stream\ndata: {}\n\n"

    headers = {
        "Content-Type": "text/event-stream",
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no"
    }
    return Response(stream_with_context(generate_and_update_memory()), headers=headers)


@chat_bp.route('/memory-stats/<session_id>', methods=['GET'])
@optional_token_required
def get_memory_stats(current_user, session_id):
    """Debug endpoint to view memory statistics."""
    if not current_user:
        return jsonify({"error": "Authentication required"}), 401

    try:
        memory = TokenAwareMemoryManager(current_user['id'], session_id)
        stats = memory.get_memory_stats()
        return jsonify(stats)
    except Exception as e:
        logging.error(f"Failed to get memory stats: {e}", exc_info=True)
        return jsonify({"error": "Failed to retrieve memory stats"}), 500