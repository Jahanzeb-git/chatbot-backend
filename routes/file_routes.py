import os
import io
import logging
import uuid
from datetime import datetime, timezone
from flask import Blueprint, request, jsonify, current_app
from werkzeug.utils import secure_filename
from auth import token_required
from db import get_db_connection, return_db_connection
import boto3
from botocore.client import Config
from botocore.exceptions import ClientError

file_bp = Blueprint('file_bp', __name__)

# Graceful imports with fallbacks
try:
    import magic
    HAS_MAGIC = True
except ImportError:
    HAS_MAGIC = False
    logging.warning("python-magic not available. File type detection will be limited.")

try:
    import pypdf
    HAS_PYPDF = True
except ImportError:
    HAS_PYPDF = False
    logging.warning("pypdf not available. PDF processing will be disabled.")

try:
    import docx
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False
    logging.warning("python-docx not available. DOCX processing will be disabled.")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logging.warning("openpyxl not available. XLSX processing will be disabled.")

# MIME type mappings (same as before)
MIME_TYPE_MAP = {
    'text/plain': '.txt',
    'application/pdf': '.pdf',
    'application/msword': '.doc',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document': '.docx',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': '.xlsx',
    'text/markdown': '.md',
    'text/csv': '.csv',
    'application/json': '.json',
    'application/xml': '.xml',
    'text/html': '.html',
    'text/css': '.css',
    'application/javascript': '.js',
    'application/x-python-code': '.py',
    'text/x-python': '.py',
    'text/x-c': '.c',
    'text/x-c++src': '.cpp',
    'text/x-java-source': '.java',
    'application/x-sh': '.sh',
    'image/jpeg': '.jpeg',
    'image/png': '.png',
}

EXT_TO_MIME = {
    '.txt': 'text/plain',
    '.pdf': 'application/pdf',
    '.doc': 'application/msword',
    '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    '.md': 'text/markdown',
    '.csv': 'text/csv',
    '.json': 'application/json',
    '.xml': 'application/xml',
    '.html': 'text/html',
    '.css': 'text/css',
    '.js': 'application/javascript',
    '.py': 'text/x-python',
    '.c': 'text/x-c',
    '.cpp': 'text/x-c++src',
    '.java': 'text/x-java-source',
    '.sh': 'application/x-sh',
    '.jpg': 'image/jpeg',
    '.jpeg': 'image/jpeg',
    '.png': 'image/png',
}

def get_b2_client():
    """Create and return a configured B2 S3 client."""
    return boto3.client(
        "s3",
        endpoint_url=current_app.config['B2_ENDPOINT'],
        aws_access_key_id=current_app.config['B2_KEY_ID'],
        aws_secret_access_key=current_app.config['B2_APP_KEY'],
        config=Config(signature_version="s3v4"),
    )

def get_file_extension(mime_type):
    """Get the file extension from the MIME type."""
    return MIME_TYPE_MAP.get(mime_type)

def detect_mime_type(file_content_bytes, filename):
    """Detect MIME type with fallback methods."""
    if HAS_MAGIC:
        try:
            return magic.from_buffer(file_content_bytes, mime=True)
        except Exception as e:
            logging.warning(f"Magic detection failed: {e}")

    if filename:
        _, ext = os.path.splitext(filename.lower())
        if ext in EXT_TO_MIME:
            return EXT_TO_MIME[ext]

    try:
        file_content_bytes.decode('utf-8')
        return 'text/plain'
    except UnicodeDecodeError:
        pass

    if file_content_bytes.startswith(b'%PDF'):
        return 'application/pdf'

    if file_content_bytes.startswith(b'PK'):
        if filename:
            if filename.lower().endswith('.docx'):
                return 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            elif filename.lower().endswith('.xlsx'):
                return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return 'application/octet-stream'

def get_user_file_count(user_id):
    """Get total number of files stored for a user."""
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT COUNT(*) as count FROM uploaded_files WHERE user_id = %s",
            (user_id,)
        )
        result = cursor.fetchone()
        return result['count'] if result else 0
    finally:
        return_db_connection(conn)

def format_file_size(size_bytes):
    """Format file size in human-readable format."""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} MB"

def generate_presigned_url(b2_key, expiration=3600):
    """Generate a presigned URL for file access."""
    try:
        s3_client = get_b2_client()
        url = s3_client.generate_presigned_url(
            ClientMethod='get_object',
            Params={
                'Bucket': current_app.config['B2_BUCKET_NAME'],
                'Key': b2_key
            },
            ExpiresIn=expiration
        )
        return url
    except Exception as e:
        logging.error(f"Failed to generate presigned URL: {e}", exc_info=True)
        return None

def extract_text_from_pdf(file_content_bytes):
    """Extract text from PDF bytes."""
    try:
        import pypdf
        pdf_file = io.BytesIO(file_content_bytes)
        pdf_reader = pypdf.PdfReader(pdf_file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text()
        return text if text.strip() else "[PDF content could not be extracted]"
    except Exception as e:
        logging.warning(f"PDF extraction failed: {e}")
        return f"[PDF content extraction error: {str(e)}]"

def extract_text_from_docx(file_content_bytes):
    """Extract text from DOCX bytes."""
    try:
        import docx
        docx_file = io.BytesIO(file_content_bytes)
        doc = docx.Document(docx_file)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text if text.strip() else "[DOCX content could not be extracted]"
    except Exception as e:
        logging.warning(f"DOCX extraction failed: {e}")
        return f"[DOCX content extraction error: {str(e)}]"

def extract_text_from_xlsx(file_content_bytes):
    """Extract text from XLSX bytes."""
    try:
        import openpyxl
        xlsx_file = io.BytesIO(file_content_bytes)
        workbook = openpyxl.load_workbook(xlsx_file)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = ",".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
            text += "\n"
        return text if text.strip() else "[XLSX content could not be extracted]"
    except Exception as e:
        logging.warning(f"XLSX extraction failed: {e}")
        return f"[XLSX content extraction error: {str(e)}]"

def extract_file_content_from_bytes(file_content_bytes, mime_type):
    """Extract content from file bytes based on type."""
    try:
        if mime_type == 'application/pdf':
            return extract_text_from_pdf(file_content_bytes)
        elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
            return extract_text_from_docx(file_content_bytes)
        elif mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            return extract_text_from_xlsx(file_content_bytes)
        else:
            # Try to decode as text
            return file_content_bytes.decode('utf-8')
    except UnicodeDecodeError:
        # Try other encodings
        for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
            try:
                return file_content_bytes.decode(encoding)
            except:
                continue
        return "[Binary file - content not readable]"
    except Exception as e:
        logging.error(f"File content extraction error: {e}")
        return f"[Error reading file: {str(e)}]"

@file_bp.route('/upload', methods=['POST'])
@token_required
def upload_files(current_user):
    """Upload 1-5 files for a chat session to Backblaze B2."""
    try:
        user_id = current_user['id']
        session_id = request.form.get('session_id')

        if not session_id:
            return jsonify({"error": "session_id is required"}), 400

        # Get uploaded files
        files = request.files.getlist('files')

        if not files or len(files) == 0:
            return jsonify({"error": "No files provided"}), 400

        if len(files) > current_app.config['MAX_FILES_PER_PROMPT']:
            return jsonify({
                "error": f"Maximum {current_app.config['MAX_FILES_PER_PROMPT']} files per prompt"
            }), 400

        # Check user's total file count
        current_file_count = get_user_file_count(user_id)
        if current_file_count + len(files) > current_app.config['MAX_FILES_PER_USER']:
            return jsonify({
                "error": f"File limit exceeded. Maximum {current_app.config['MAX_FILES_PER_USER']} files. Current: {current_file_count}"
            }), 400

        s3_client = get_b2_client()
        uploaded_files_metadata = []
        total_size = 0
        file_ids = []

        conn = get_db_connection()

        try:
            for file in files:
                if not file or file.filename == '':
                    continue

                # Read file content
                file_content_bytes = file.read()
                file_size = len(file_content_bytes)

                if file_size == 0:
                    return jsonify({"error": f"File {file.filename} is empty"}), 400

                if file_size > current_app.config['MAX_FILE_SIZE_BYTES']:
                    return jsonify({
                        "error": f"File {file.filename} exceeds 10MB limit"
                    }), 400

                # Detect MIME type
                mime_type = detect_mime_type(file_content_bytes, file.filename)

                # Generate B2 key
                file_ext = get_file_extension(mime_type) or os.path.splitext(file.filename)[1] or '.bin'
                b2_key = f"user_uploads/{user_id}/{uuid.uuid4()}{file_ext}"

                # Upload to B2
                s3_client.put_object(
                    Bucket=current_app.config['B2_BUCKET_NAME'],
                    Key=b2_key,
                    Body=file_content_bytes,
                    ContentType=mime_type
                )

                is_image = 1 if mime_type in ['image/jpeg', 'image/png'] else 0
                uploaded_at = datetime.now(timezone.utc).isoformat()

                # Insert into database
                cursor = conn.cursor()
                cursor.execute(
                    """INSERT INTO uploaded_files
                       (user_id, session_number, b2_key, original_name, size, mime_type, is_image, uploaded_at)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                       RETURNING id""",
                    (user_id, int(session_id), b2_key, file.filename, file_size, mime_type, is_image, uploaded_at)
                )
                file_id = cursor.fetchone()['id']
                file_ids.append(file_id)

                uploaded_files_metadata.append({
                    "b2_key": b2_key,
                    "original_name": file.filename,
                    "size": file_size,
                    "type": mime_type,
                    "is_image": bool(is_image)
                })

                total_size += file_size
                logging.info(f"Successfully uploaded {file.filename} to B2: {b2_key}")

            conn.commit()

            # Stage file IDs in cache for the next chat request
            cache_key = f"{user_id}-{session_id}"
            if not hasattr(current_app, 'file_cache'):
                current_app.file_cache = {}
            current_app.file_cache[cache_key] = file_ids

            return jsonify({
                "message": f"{len(uploaded_files_metadata)} file(s) staged successfully",
                "files": uploaded_files_metadata,
                "total_size": total_size
            }), 200

        except ClientError as e:
            conn.rollback()
            logging.error(f"B2 upload error: {e}", exc_info=True)
            return jsonify({"error": f"File upload to B2 failed: {str(e)}"}), 500
        except Exception as e:
            conn.rollback()
            logging.error(f"File upload error: {e}", exc_info=True)
            return jsonify({"error": f"File upload failed: {str(e)}"}), 500
        finally:
            return_db_connection(conn)

    except Exception as e:
        logging.error(f"File upload error: {e}", exc_info=True)
        return jsonify({"error": f"File upload failed: {str(e)}"}), 500

@file_bp.route('/files/<session_id>/<path:b2_key>', methods=['GET'])
@token_required
def get_file_content(current_user, session_id, b2_key):
    """Get file content by generating presigned URL."""
    user_id = current_user['id']
    conn = get_db_connection()

    try:
        # Verify file belongs to user and session
        cursor = conn.cursor()
        cursor.execute(
            """SELECT b2_key, original_name, mime_type, is_image
               FROM uploaded_files
               WHERE user_id = %s AND session_number = %s AND b2_key = %s""",
            (user_id, int(session_id), b2_key)
        )
        file_record = cursor.fetchone()

        if not file_record:
            return jsonify({"error": "File not found or access denied"}), 404

        # Generate presigned URL
        presigned_url = generate_presigned_url(b2_key)
        
        if not presigned_url:
            return jsonify({"error": "Failed to generate file URL"}), 500

        return jsonify({
            "url": presigned_url,
            "original_name": file_record['original_name'],
            "mime_type": file_record['mime_type']
        }), 200

    except Exception as e:
        logging.error(f"Error retrieving file: {e}", exc_info=True)
        return jsonify({"error": "Failed to retrieve file"}), 500
    finally:
        return_db_connection(conn)

@file_bp.route('/files/list', methods=['GET'])
@token_required
def list_user_files(current_user):
    """List all files for the user, optionally filtered by session."""
    user_id = current_user['id']
    session_number = request.args.get('session_number', type=int)

    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        if session_number:
            cursor.execute(
                """SELECT id, b2_key, original_name, size, mime_type, is_image, uploaded_at, session_number
                   FROM uploaded_files
                   WHERE user_id = %s AND session_number = %s
                   ORDER BY uploaded_at DESC""",
                (user_id, session_number)
            )
        else:
            cursor.execute(
                """SELECT id, b2_key, original_name, size, mime_type, is_image, uploaded_at, session_number
                   FROM uploaded_files
                   WHERE user_id = %s
                   ORDER BY uploaded_at DESC""",
                (user_id,)
            )
        
        files = cursor.fetchall()
        files_list = [dict(file) for file in files]
        
        return jsonify({
            "files": files_list,
            "total": len(files_list)
        }), 200

    except Exception as e:
        logging.error(f"Error listing files: {e}", exc_info=True)
        return jsonify({"error": "Failed to list files"}), 500
    finally:
        return_db_connection(conn)

@file_bp.route('/files', methods=['DELETE'])
@token_required
def delete_files(current_user):
    """Delete specific files or all files for the user from B2 and database."""
    user_id = current_user['id']
    data = request.json or {}

    delete_all = data.get('delete_all', False)
    b2_keys = data.get('b2_keys', [])

    if not delete_all and not b2_keys:
        return jsonify({"error": "Provide 'b2_keys' array or 'delete_all': true"}), 400

    conn = get_db_connection()
    s3_client = get_b2_client()

    try:
        cursor = conn.cursor()
        
        if delete_all:
            # Get all user's files
            cursor.execute(
                "SELECT b2_key FROM uploaded_files WHERE user_id = %s",
                (user_id,)
            )
            files = cursor.fetchall()

            # Delete from B2
            for file in files:
                try:
                    s3_client.delete_object(
                        Bucket=current_app.config['B2_BUCKET_NAME'],
                        Key=file['b2_key']
                    )
                    logging.info(f"Deleted from B2: {file['b2_key']}")
                except Exception as e:
                    logging.warning(f"Failed to delete from B2 {file['b2_key']}: {e}")

            # Delete from database
            cursor.execute("DELETE FROM uploaded_files WHERE user_id = %s", (user_id,))
            conn.commit()

            return jsonify({"message": f"Deleted {len(files)} file(s)"}), 200

        else:
            # Delete specific files
            deleted_count = 0
            for b2_key in b2_keys:
                # Verify ownership
                cursor.execute(
                    "SELECT id FROM uploaded_files WHERE user_id = %s AND b2_key = %s",
                    (user_id, b2_key)
                )
                file_record = cursor.fetchone()

                if file_record:
                    # Delete from B2
                    try:
                        s3_client.delete_object(
                            Bucket=current_app.config['B2_BUCKET_NAME'],
                            Key=b2_key
                        )
                        logging.info(f"Deleted from B2: {b2_key}")
                    except Exception as e:
                        logging.warning(f"Failed to delete from B2 {b2_key}: {e}")

                    # Delete from database
                    cursor.execute(
                        "DELETE FROM uploaded_files WHERE user_id = %s AND b2_key = %s",
                        (user_id, b2_key)
                    )
                    deleted_count += 1

            conn.commit()
            return jsonify({"message": f"Deleted {deleted_count} file(s)"}), 200

    except Exception as e:
        conn.rollback()
        logging.error(f"Error deleting files: {e}", exc_info=True)
        return jsonify({"error": "Failed to delete files"}), 500
    finally:
        return_db_connection(conn)

@file_bp.route('/upload/status', methods=['GET'])
@token_required
def upload_status(current_user):
    """Check if there are staged files for the session."""
    session_id = request.args.get('session_id')
    if not session_id:
        return jsonify({"error": "session_id is required"}), 400

    user_id = current_user['id']
    cache_key = f"{user_id}-{session_id}"

    if hasattr(current_app, 'file_cache') and cache_key in current_app.file_cache:
        file_ids = current_app.file_cache[cache_key]

        # Get metadata for staged files
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            placeholders = ','.join(['%s'] * len(file_ids))
            cursor.execute(
                f"""SELECT b2_key, original_name, size, mime_type
                   FROM uploaded_files
                   WHERE id IN ({placeholders})""",
                file_ids
            )
            files = cursor.fetchall()

            return jsonify({
                "has_files": True,
                "files": [dict(f) for f in files],
                "count": len(files)
            }), 200
        finally:
            return_db_connection(conn)

    return jsonify({"has_files": False, "count": 0}), 200

@file_bp.route('/upload/clear', methods=['POST'])
@token_required
def clear_upload(current_user):
    """Clear staged files for the session."""
    data = request.json or {}
    session_id = data.get('session_id')
    if not session_id:
        return jsonify({"error": "session_id is required"}), 400

    user_id = current_user['id']
    cache_key = f"{user_id}-{session_id}"

    if hasattr(current_app, 'file_cache') and cache_key in current_app.file_cache:
        del current_app.file_cache[cache_key]
        return jsonify({"message": "Staged files cleared successfully"}), 200

    return jsonify({"message": "No files to clear"}), 200